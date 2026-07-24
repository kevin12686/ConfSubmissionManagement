from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="26364A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_BORDER = Border(bottom=Side(style="medium", color="52657A"))
ROW_BORDER = Border(bottom=Side(style="thin", color="9AA8B8"))

LONG_TEXT_TOKENS = {
    "author",
    "blocking",
    "exception",
    "message",
    "note",
    "reason",
    "title",
}
IDENTIFIER_HEADERS = {
    "author entered paper id",
    "final id",
    "final submission id",
    "final_submission_id",
    "id",
    "paper id",
    "paper_id",
    "paper_id_filled",
    "start2_paper_id_raw",
}
PERCENT_HEADERS = {
    "plagiarism %",
    "similarity (p)",
    "similarity (s)",
    "similarity_score",
    "single %",
    "single_similarity_score",
}


def _header_key(value):
    return str(value or "").strip().casefold()


def _display_length(value):
    if value is None:
        return 0
    return max((len(line) for line in str(value).splitlines()), default=0)


def _column_width(header, values):
    key = _header_key(header)
    content_width = max(
        [len(str(header))] + [_display_length(value) for value in values],
        default=len(str(header)),
    )
    if "title" in key:
        return min(max(content_width + 2, 24), 52)
    if any(token in key for token in LONG_TEXT_TOKENS):
        return min(max(content_width + 2, 18), 46)
    if "path" in key or "file" in key or "image" in key:
        return min(max(content_width + 2, 20), 42)
    if key in IDENTIFIER_HEADERS or key.endswith("_id"):
        return min(max(content_width + 2, 14), 24)
    if key.endswith("_at") or "date" in key or "time" in key:
        return min(max(content_width + 2, 18), 22)
    return min(max(content_width + 2, 12), 30)


def _is_long_text_header(header):
    key = _header_key(header)
    return any(token in key for token in LONG_TEXT_TOKENS)


def _format_worksheet(worksheet):
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.row_dimensions[1].height = 26

    if worksheet.max_column < 1:
        return

    headers = [
        worksheet.cell(row=1, column=column).value or ""
        for column in range(1, worksheet.max_column + 1)
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        values = [
            worksheet.cell(row=row, column=column).value
            for row in range(2, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column)].width = _column_width(
            header,
            values,
        )

        key = _header_key(header)
        for row in range(2, worksheet.max_row + 1):
            data_cell = worksheet.cell(row=row, column=column)
            data_cell.border = ROW_BORDER
            data_cell.alignment = Alignment(
                vertical="top",
                wrap_text=_is_long_text_header(header),
            )
            if key in IDENTIFIER_HEADERS or key.endswith("_id"):
                data_cell.number_format = "@"
            elif key in PERCENT_HEADERS:
                data_cell.number_format = '0"%"'
            elif key.endswith("_at") or "date" in key or "time" in key:
                data_cell.number_format = "yyyy-mm-dd hh:mm"

    if any(headers):
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}"
            f"{max(worksheet.max_row, 1)}"
        )


def write_formatted_workbook(path, sheets):
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_worksheet(writer.book[sheet_name])
        writer.book.active = 0
    return path
